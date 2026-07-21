"""Ground truth: body-mass change per subject/session.

Body mass is the ONLY objective hydration reference in this study. The signed target is

    Delta m%(subject, session) = (m(session) - m(S0)) / m(S0) * 100      (negative = loss)

Parsing is by FIXED CELL ADDRESS, never by header-name inference: the sheet has a
two-row merged header (row 1 column labels, row 2 fractional-time subheaders) that
pandas/openpyxl header inference would mangle. The sheet also reports ~1000 rows x 113
columns because of stray formatting, so max_row/max_column must never be used to
establish the cohort size.

Module structure exists for testability. openpyxl writes formulas but never evaluates
them, so a synthetic workbook can hold EITHER a formula in column J OR a cached number,
never both — meaning no synthetic fixture can exercise the full dual-view load. The
work is therefore split into three independently testable pieces:

    _validate_layout(ws_formula)   headers, subject identity, J formula structure
    _read_values(ws_data_only)     masses, covariates, cached J, K text
    check_targets(...)             pure math + both cross-checks, no I/O

with load_ground_truth() the only place the two worksheet views meet (exercised on the
real workbook in a realdata test).
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from .sessions import SESSION_NAMES

SHEET = "MetaData"

FIRST_ROW = 3
LAST_ROW = 18
N_SUBJECTS = LAST_ROW - FIRST_ROW + 1

COL_NAME = 2        # B
COL_AGE = 3         # C
COL_HEIGHT = 4      # D
COL_WEIGHT_FIRST = 5   # E = 8am ... I = 4pm
COL_WEIGHT_LAST = 9
COL_KG_LOST = 10    # J  (an =I-E formula; cached value read via data_only)
COL_NOTE = 11       # K  ("Loss of 1.74% of body weight")

# Expected header text. Row 2 mixes datetime.time cells with one literal string.
EXPECTED_ROW1 = {
    COL_NAME: "Name",
    COL_AGE: "Age",
    COL_HEIGHT: "Height (cm)",
    COL_WEIGHT_FIRST: "Weight (kg)",
    COL_KG_LOST: "Weight lost (kg)",
    COL_NOTE: "Observations",
}
EXPECTED_ROW2_TIMES = {5: (8, 0), 6: (10, 0), 8: (14, 0), 9: (16, 0)}
EXPECTED_ROW2_NOON = (7, "12 Noon")

# Cross-check tolerances. These are conservative bounds justified by direct inspection
# of the real workbook (see HISTORY.md), NOT claims about uniform recording precision:
# most weights step by 0.1 kg but Subject 15 uses 0.05 kg increments, and column K is
# not always nearest-two-decimal rounding (Subject 13 shows truncation). The worst
# observed K deviation is ~0.0097 pct-points, so 0.05 is ~5x the observed worst case.
TOL_KG = 0.05
TOL_PCT = 0.05

# Generous plausibility bands: these catch wrong-cell parsing, not unusual subjects.
MASS_RANGE_KG = (30.0, 200.0)
AGE_RANGE_YR = (15, 80)
HEIGHT_RANGE_CM = (120.0, 220.0)
DELTA_PCT_RANGE = (-10.0, 5.0)

_SUBJECT_RE = re.compile(r"^Subject\s+(\d+)$")
_NUMBER_RE = re.compile(r"(\d+(?:\.\d+)?)")


class GroundTruthError(ValueError):
    """Raised when the workbook's layout, identity, or arithmetic fails validation."""


@dataclass(frozen=True)
class GroundTruth:
    sessions: pd.DataFrame  # subject, session_idx, session_name, mass_kg,
                            # delta_m_kg, delta_m_pct
    subjects: pd.DataFrame  # subject, age, height_cm, baseline_mass_kg, bmi


# ----------------------------------------------------------------- layout validation


def _validate_layout(ws) -> dict[int, int]:
    """Validate headers, subject identity and J's formula on the FORMULA view.

    Returns {row -> subject_id}. Never reads cached values, so it is unit-testable
    with an openpyxl-written fixture.
    """
    problems: list[str] = []

    for col, expected in EXPECTED_ROW1.items():
        actual = ws.cell(row=1, column=col).value
        if actual != expected:
            problems.append(f"header row1 col{col}: expected {expected!r}, got {actual!r}")

    for col, (hour, minute) in EXPECTED_ROW2_TIMES.items():
        actual = ws.cell(row=2, column=col).value
        if not (hasattr(actual, "hour") and (actual.hour, actual.minute) == (hour, minute)):
            problems.append(
                f"header row2 col{col}: expected time {hour:02d}:{minute:02d}, got {actual!r}"
            )
    noon_col, noon_text = EXPECTED_ROW2_NOON
    if ws.cell(row=2, column=noon_col).value != noon_text:
        problems.append(
            f"header row2 col{noon_col}: expected {noon_text!r}, "
            f"got {ws.cell(row=2, column=noon_col).value!r}"
        )

    # Subject identity comes from column B, not from row position.
    row_to_subject: dict[int, int] = {}
    for row in range(FIRST_ROW, LAST_ROW + 1):
        raw = ws.cell(row=row, column=COL_NAME).value
        match = _SUBJECT_RE.match(str(raw).strip()) if raw is not None else None
        if match is None:
            problems.append(f"row {row}: column B is {raw!r}, expected 'Subject <id>'")
            continue
        row_to_subject[row] = int(match.group(1))

    ids = sorted(row_to_subject.values())
    if ids and ids != list(range(1, N_SUBJECTS + 1)):
        problems.append(f"subject ids are {ids}, expected 1..{N_SUBJECTS} exactly once each")

    # A subject added ANYWHERE else in column B must be caught, not just below the block.
    for row in range(1, ws.max_row + 1):
        if FIRST_ROW <= row <= LAST_ROW:
            continue
        raw = ws.cell(row=row, column=COL_NAME).value
        if raw is not None and _SUBJECT_RE.match(str(raw).strip()):
            problems.append(f"unexpected subject record outside rows {FIRST_ROW}-{LAST_ROW}: B{row}={raw!r}")

    # Column J must be the =I-E difference, not a hand-typed number.
    for row in range(FIRST_ROW, LAST_ROW + 1):
        formula = ws.cell(row=row, column=COL_KG_LOST).value
        expected = f"=I{row}-E{row}"
        if formula != expected:
            problems.append(f"J{row}: expected formula {expected!r}, got {formula!r}")

    if problems:
        raise GroundTruthError(
            "workbook layout validation failed:\n  " + "\n  ".join(problems)
        )
    return row_to_subject


# --------------------------------------------------------------------- value reading


def _is_plausible_mass(value) -> bool:
    """A weight cell must be a finite number inside a generous plausibility band."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    return math.isfinite(value) and MASS_RANGE_KG[0] <= value <= MASS_RANGE_KG[1]


def _read_values(ws, row_to_subject: dict[int, int]) -> list[dict]:
    """Read masses, covariates, cached J and K text from the DATA_ONLY view."""
    problems: list[str] = []
    records: list[dict] = []

    for row, subject in sorted(row_to_subject.items()):
        masses = [
            ws.cell(row=row, column=col).value
            for col in range(COL_WEIGHT_FIRST, COL_WEIGHT_LAST + 1)
        ]
        age = ws.cell(row=row, column=COL_AGE).value
        height = ws.cell(row=row, column=COL_HEIGHT).value
        kg_lost = ws.cell(row=row, column=COL_KG_LOST).value
        note = ws.cell(row=row, column=COL_NOTE).value

        bad_mass = [
            (SESSION_NAMES[i], m)
            for i, m in enumerate(masses)
            if not _is_plausible_mass(m)
        ]
        if bad_mass:
            problems.append(f"subject {subject}: implausible/missing weights {bad_mass}")
            continue

        # Covariates are validated BEFORE BMI is computed from them.
        if not isinstance(age, (int, float)) or not (AGE_RANGE_YR[0] <= age <= AGE_RANGE_YR[1]):
            problems.append(f"subject {subject}: implausible age {age!r}")
            continue
        if not isinstance(height, (int, float)) or not (
            HEIGHT_RANGE_CM[0] <= height <= HEIGHT_RANGE_CM[1]
        ):
            problems.append(f"subject {subject}: implausible height {height!r}")
            continue
        if not isinstance(kg_lost, (int, float)):
            problems.append(
                f"subject {subject}: column J has no cached numeric value ({kg_lost!r}); "
                "the workbook must be saved by Excel so the formula result is stored"
            )
            continue

        records.append(
            {
                "subject": subject,
                "age": int(age),
                "height_cm": float(height),
                "masses": [float(m) for m in masses],
                "kg_lost": float(kg_lost),
                "note": note,
            }
        )

    if problems:
        raise GroundTruthError("workbook value validation failed:\n  " + "\n  ".join(problems))
    return records


# ------------------------------------------------------------------ targets & checks


def parse_note_percentage(note) -> float | None:
    """Extract the positive percentage from e.g. 'Loss of 1.74% of body weight'."""
    if not isinstance(note, str):
        return None
    match = _NUMBER_RE.search(note)
    return float(match.group(1)) if match else None


def check_targets(subject: int, masses: list[float], kg_lost: float, note) -> list[str]:
    """Compute nothing, verify everything: the two sign-aware cross-checks.

    Pure function (no I/O), so tolerance behaviour and every failure mode are testable
    directly with arrays. Returns a list of human-readable discrepancies.
    """
    problems: list[str] = []
    baseline = masses[0]
    delta_kg_s4 = masses[-1] - baseline
    delta_pct_s4 = delta_kg_s4 / baseline * 100.0

    # (i) column J is a signed kg change.
    if abs(delta_kg_s4 - kg_lost) > TOL_KG:
        problems.append(
            f"subject {subject}: computed m(S4)-m(S0) = {delta_kg_s4:.3f} kg but "
            f"column J states {kg_lost:.3f} kg (tolerance {TOL_KG} kg)"
        )

    # (ii) column K is a POSITIVE percentage in text.
    stated_pct = parse_note_percentage(note)
    if stated_pct is None:
        problems.append(f"subject {subject}: cannot parse a percentage from note {note!r}")
    elif abs(stated_pct - abs(delta_pct_s4)) > TOL_PCT:
        problems.append(
            f"subject {subject}: computed |Delta m%(S4)| = {abs(delta_pct_s4):.4f}% but "
            f"column K states {stated_pct:.4f}% (tolerance {TOL_PCT} pct-points)"
        )

    for idx, mass in enumerate(masses):
        delta_pct = (mass - baseline) / baseline * 100.0
        if not (DELTA_PCT_RANGE[0] <= delta_pct <= DELTA_PCT_RANGE[1]):
            problems.append(
                f"subject {subject}, {SESSION_NAMES[idx]}: Delta m% = {delta_pct:.3f} "
                f"outside plausible range {DELTA_PCT_RANGE}"
            )
    return problems


# ------------------------------------------------------------------------ public API


def load_ground_truth(xlsx_path: str | Path) -> GroundTruth:
    """Parse the workbook into per-session targets and per-subject covariates.

    Fails loudly, listing every offending subject, on any layout, identity, value or
    cross-check problem. A genuine cross-check failure is investigated — the tolerance
    is never widened to make it pass.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.is_file():
        raise GroundTruthError(f"ground-truth workbook not found: {xlsx_path}")

    # Two views of the same file: formulas cannot be read from the data_only view, and
    # cached values cannot be read from the formula view.
    wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    for wb in (wb_formula, wb_values):
        if SHEET not in wb.sheetnames:
            raise GroundTruthError(f"{xlsx_path}: no sheet named {SHEET!r}")

    row_to_subject = _validate_layout(wb_formula[SHEET])
    records = _read_values(wb_values[SHEET], row_to_subject)

    if len(records) != N_SUBJECTS:
        raise GroundTruthError(f"parsed {len(records)} subjects, expected {N_SUBJECTS}")

    problems: list[str] = []
    for record in records:
        problems.extend(
            check_targets(record["subject"], record["masses"], record["kg_lost"], record["note"])
        )
    if problems:
        raise GroundTruthError(
            "ground-truth cross-checks failed:\n  " + "\n  ".join(problems)
        )

    session_rows = []
    subject_rows = []
    for record in records:
        masses = record["masses"]
        baseline = masses[0]
        for idx, mass in enumerate(masses):
            session_rows.append(
                {
                    "subject": record["subject"],
                    "session_idx": idx,
                    "session_name": SESSION_NAMES[idx],
                    "mass_kg": mass,
                    "delta_m_kg": mass - baseline,
                    "delta_m_pct": (mass - baseline) / baseline * 100.0,
                }
            )
        height_m = record["height_cm"] / 100.0
        subject_rows.append(
            {
                "subject": record["subject"],
                "age": record["age"],
                "height_cm": record["height_cm"],
                "baseline_mass_kg": baseline,
                "bmi": baseline / (height_m**2),
            }
        )

    sessions = pd.DataFrame(session_rows).sort_values(["subject", "session_idx"])
    subjects = pd.DataFrame(subject_rows).sort_values("subject")
    return GroundTruth(
        sessions=sessions.reset_index(drop=True),
        subjects=subjects.reset_index(drop=True),
    )
