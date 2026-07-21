"""Ground-truth parsing, layout validation and the two sign-aware cross-checks.

Test strategy follows from an openpyxl limitation (verified): openpyxl writes formulas
but never evaluates them, so reopening an openpyxl-written file with data_only=True
returns None for a formula cell. A synthetic workbook can therefore hold EITHER a
formula in column J OR a cached number — never both — and no synthetic fixture can
exercise the full dual-view load_ground_truth(). Hence:

  * check_targets      -> pure-function tests, no workbook at all
  * _validate_layout   -> formula-view fixtures (formulas written as strings)
  * _read_values       -> literal-value fixtures (J written as a number)
  * load_ground_truth  -> realdata test only (the real file has formulas AND caches)
"""

import datetime as dt

import openpyxl
import pytest

from dehyd.data.ground_truth import (
    COL_KG_LOST,
    FIRST_ROW,
    LAST_ROW,
    SHEET,
    GroundTruthError,
    _read_values,
    _validate_layout,
    check_targets,
    load_ground_truth,
    parse_note_percentage,
)

# Realistic per-subject rows: (age, height_cm, masses E..I, kg_lost, note)
GOOD_ROWS = [
    (27, 187, [86.3, 85.8, 85.6, 85.1, 84.8], -1.5, "Loss of 1.74% of body weight"),
    (28, 188, [98.3, 98.2, 98.1, 97.5, 97.2], -1.1, "Loss of 1.12% of body weight"),
]


def build_sheet(ws, rows=GOOD_ROWS, *, j_as_formula: bool, n=None):
    """Write a workbook with the real layout. j_as_formula picks which view works."""
    ws["B1"], ws["C1"], ws["D1"] = "Name", "Age", "Height (cm)"
    ws["E1"], ws["J1"], ws["K1"] = "Weight (kg)", "Weight lost (kg)", "Observations"
    ws["E2"], ws["F2"] = dt.time(8, 0), dt.time(10, 0)
    ws["G2"] = "12 Noon"
    ws["H2"], ws["I2"] = dt.time(14, 0), dt.time(16, 0)

    n = n if n is not None else len(rows)
    for i in range(n):
        age, height, masses, kg_lost, note = rows[i % len(rows)]
        row = FIRST_ROW + i
        ws.cell(row=row, column=2, value=f"Subject {i + 1}")
        ws.cell(row=row, column=3, value=age)
        ws.cell(row=row, column=4, value=height)
        for k, mass in enumerate(masses):
            ws.cell(row=row, column=5 + k, value=mass)
        ws.cell(
            row=row,
            column=COL_KG_LOST,
            value=f"=I{row}-E{row}" if j_as_formula else kg_lost,
        )
        ws.cell(row=row, column=11, value=note)
    return ws


def make_ws(*, j_as_formula=True, rows=GOOD_ROWS, n=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    return build_sheet(ws, rows, j_as_formula=j_as_formula, n=n)


def full_row_map(n=16):
    return {FIRST_ROW + i: i + 1 for i in range(n)}


# =============================================================== check_targets (pure)


def test_cross_checks_pass_on_consistent_values():
    assert check_targets(1, [86.3, 85.8, 85.6, 85.1, 84.8], -1.5, "Loss of 1.74%") == []


def test_column_j_mismatch_detected():
    problems = check_targets(1, [86.3, 85.8, 85.6, 85.1, 84.8], -0.9, "Loss of 1.74%")
    assert any("column J" in p for p in problems)


def test_column_k_mismatch_detected():
    problems = check_targets(1, [86.3, 85.8, 85.6, 85.1, 84.8], -1.5, "Loss of 0.90%")
    assert any("column K" in p for p in problems)


def test_unparseable_note_detected():
    problems = check_targets(1, [86.3, 85.8, 85.6, 85.1, 84.8], -1.5, None)
    assert any("cannot parse" in p for p in problems)


def test_tolerances_are_inclusive_at_the_boundary():
    """0.05 kg / 0.05 pct-points are conservative bounds, not exact-equality demands."""
    masses = [100.0, 99.9, 99.8, 99.7, 99.0]  # exactly -1.0 kg, -1.00%
    assert check_targets(1, masses, -1.04, "Loss of 1.00%") == []      # within 0.05 kg
    assert check_targets(1, masses, -1.0, "Loss of 1.04%") == []       # within 0.05 pct
    assert check_targets(1, masses, -1.06, "Loss of 1.00%") != []      # outside
    assert check_targets(1, masses, -1.0, "Loss of 1.20%") != []       # outside


def test_sign_convention_is_negative_for_loss():
    """Delta m% must be negative when mass falls; K states the positive magnitude."""
    problems = check_targets(1, [100.0, 99.9, 99.8, 99.7, 99.0], +1.0, "Loss of 1.00%")
    assert any("column J" in p for p in problems)  # +1.0 has the wrong sign


def test_implausible_delta_detected():
    problems = check_targets(1, [100.0, 99.0, 98.0, 95.0, 80.0], -20.0, "Loss of 20.00%")
    assert any("outside plausible range" in p for p in problems)


@pytest.mark.parametrize(
    "note,expected",
    [
        ("Loss of 1.74% of body weight", 1.74),
        ("Loss of 2% of body weight", 2.0),
        ("no numbers here", None),
        (None, None),
        (123, None),
    ],
)
def test_parse_note_percentage(note, expected):
    assert parse_note_percentage(note) == expected


# ======================================================== _validate_layout (formulas)


def test_layout_valid():
    ws = make_ws(n=16)
    assert _validate_layout(ws) == full_row_map()


def test_bad_header_cell_detected():
    ws = make_ws(n=16)
    ws["D1"] = "Stature"
    with pytest.raises(GroundTruthError, match="header row1"):
        _validate_layout(ws)


def test_noon_header_must_be_the_literal_string():
    ws = make_ws(n=16)
    ws["G2"] = dt.time(12, 0)
    with pytest.raises(GroundTruthError, match="12 Noon"):
        _validate_layout(ws)


def test_subject_identity_from_column_b_not_row_position():
    ws = make_ws(n=16)
    ws.cell(row=FIRST_ROW, column=2, value="Participant One")
    with pytest.raises(GroundTruthError, match="expected 'Subject <id>'"):
        _validate_layout(ws)


def test_duplicate_subject_id_detected():
    ws = make_ws(n=16)
    ws.cell(row=FIRST_ROW + 1, column=2, value="Subject 1")
    with pytest.raises(GroundTruthError, match="expected 1..16"):
        _validate_layout(ws)


def test_extra_subject_row_far_below_block_detected():
    """The guard scans all of column B, not just the rows right after the block."""
    ws = make_ws(n=16)
    ws.cell(row=400, column=2, value="Subject 17")
    with pytest.raises(GroundTruthError, match="unexpected subject record"):
        _validate_layout(ws)


def test_extra_subject_row_above_block_detected():
    ws = make_ws(n=16)
    ws.cell(row=1, column=2, value="Subject 99")  # overwrites the 'Name' header
    with pytest.raises(GroundTruthError):
        _validate_layout(ws)


def test_hand_edited_column_j_detected():
    """J must be the =I-E formula, not a typed constant."""
    ws = make_ws(n=16)
    ws.cell(row=FIRST_ROW, column=COL_KG_LOST, value=-1.5)
    with pytest.raises(GroundTruthError, match="expected formula"):
        _validate_layout(ws)


def test_wrong_formula_detected():
    ws = make_ws(n=16)
    ws.cell(row=FIRST_ROW, column=COL_KG_LOST, value=f"=E{FIRST_ROW}-I{FIRST_ROW}")
    with pytest.raises(GroundTruthError, match="expected formula"):
        _validate_layout(ws)


def test_all_problems_reported_not_just_the_first():
    ws = make_ws(n=16)
    ws.cell(row=FIRST_ROW, column=COL_KG_LOST, value=-1.5)
    ws.cell(row=FIRST_ROW + 1, column=COL_KG_LOST, value=-1.1)
    with pytest.raises(GroundTruthError) as excinfo:
        _validate_layout(ws)
    assert str(excinfo.value).count("expected formula") == 2


# ========================================================== _read_values (literal J)


def test_read_values_extracts_records():
    ws = make_ws(j_as_formula=False, n=2)
    records = _read_values(ws, full_row_map(2))

    assert [r["subject"] for r in records] == [1, 2]
    assert records[0]["masses"] == GOOD_ROWS[0][2]
    assert records[0]["kg_lost"] == pytest.approx(-1.5)
    assert records[0]["age"] == 27


def test_missing_cached_j_reports_the_openpyxl_limitation():
    """A formula-only fixture has no cached value; the error must say so clearly."""
    ws = make_ws(j_as_formula=True, n=2)
    with pytest.raises(GroundTruthError, match="no cached numeric value"):
        _read_values(ws, full_row_map(2))


def test_missing_weight_detected():
    ws = make_ws(j_as_formula=False, n=2)
    # NB: ws.cell(..., value=None) is a no-op in openpyxl (None means "don't set"),
    # so the cell has to be blanked through .value directly.
    ws.cell(row=FIRST_ROW, column=7).value = None
    with pytest.raises(GroundTruthError, match="implausible/missing weights"):
        _read_values(ws, full_row_map(2))


def test_non_numeric_weight_detected():
    ws = make_ws(j_as_formula=False, n=2)
    ws.cell(row=FIRST_ROW, column=7).value = "n/a"
    with pytest.raises(GroundTruthError, match="implausible/missing weights"):
        _read_values(ws, full_row_map(2))


def test_implausible_mass_detected():
    ws = make_ws(j_as_formula=False, n=2)
    ws.cell(row=FIRST_ROW, column=5, value=8630)  # decimal point slipped
    with pytest.raises(GroundTruthError, match="implausible/missing weights"):
        _read_values(ws, full_row_map(2))


def test_implausible_age_detected_before_bmi():
    ws = make_ws(j_as_formula=False, n=2)
    ws.cell(row=FIRST_ROW, column=3, value=270)
    with pytest.raises(GroundTruthError, match="implausible age"):
        _read_values(ws, full_row_map(2))


def test_implausible_height_detected_before_bmi():
    ws = make_ws(j_as_formula=False, n=2)
    ws.cell(row=FIRST_ROW, column=4, value=1.87)  # metres instead of cm
    with pytest.raises(GroundTruthError, match="implausible height"):
        _read_values(ws, full_row_map(2))


# ============================================================== load_ground_truth (IO)


def test_missing_workbook_raises(tmp_path):
    with pytest.raises(GroundTruthError, match="not found"):
        load_ground_truth(tmp_path / "absent.xlsx")


def test_missing_sheet_raises(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    path = tmp_path / "wrong_sheet.xlsx"
    wb.save(path)
    with pytest.raises(GroundTruthError, match="no sheet named"):
        load_ground_truth(path)


# ===================================================================== realdata tests


@pytest.mark.realdata
def test_real_workbook_parses_and_cross_checks_pass(real_data_paths):
    """The only place the full dual-view path runs: real formulas AND real caches."""
    gt = load_ground_truth(real_data_paths["weight_xlsx"])

    assert len(gt.sessions) == 80
    assert len(gt.subjects) == 16
    assert sorted(gt.subjects["subject"]) == list(range(1, 17))


@pytest.mark.realdata
def test_real_targets_have_expected_structure(real_data_paths):
    gt = load_ground_truth(real_data_paths["weight_xlsx"])

    s0 = gt.sessions[gt.sessions.session_idx == 0]
    assert (s0.delta_m_pct == 0).all()          # S0 is the baseline by construction
    assert (s0.delta_m_kg == 0).all()

    s4 = gt.sessions[gt.sessions.session_idx == 4]
    assert (s4.delta_m_pct < 0).all()           # every subject lost mass by 4pm
    assert s4.delta_m_pct.min() > -3.0          # spans roughly 0 to -2%

    assert gt.subjects["bmi"].between(15, 45).all()
